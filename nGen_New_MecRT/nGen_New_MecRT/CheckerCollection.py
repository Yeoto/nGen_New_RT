import openpyxl
import re
import PyUtils
import copy
from tqdm import tqdm
import concurrent.futures

class Checker:
    def __init__(self):
        self.strID = ''
        self.list = []
        self.KeyIndex = -1
        self.bTemp = False
        self.nLine = -1
        self.bCheckedData = False

    def HasKey(self):
        return self.KeyIndex != -1

    def IsSameElement(self, DataColl_This, DataColl_Tgt, CheckData, src_data, tgt_data, bCheckerData):
        if bCheckerData:
            if type(CheckData) == type('') or type(CheckData) == type([]):
                list_RecursiveID = []
                if type(CheckData) == type(''):
                    list_RecursiveID.append(CheckData)
                elif type(CheckData) == type([]):
                    list_RecursiveID = CheckData

                bFind = False
                for strRecursiveID in list_RecursiveID:
                    if strRecursiveID not in DataColl_This.DataList or strRecursiveID not in DataColl_Tgt.DataList:
                        continue

                    RecursiveData_List_This = DataColl_This.DataList[strRecursiveID]
                    RecursiveData_List_Tgt = DataColl_Tgt.DataList[strRecursiveID]

                    if type(RecursiveData_List_This) != type({}) or type(RecursiveData_List_Tgt) != type({}):
                        continue
                    
                    if type(src_data) != type(0) or type(tgt_data) != type(0):
                        continue

                    if src_data not in RecursiveData_List_This or tgt_data not in RecursiveData_List_Tgt:
                        continue

                    src_recursive_Data = RecursiveData_List_This[src_data]
                    dst_recursive_Data = RecursiveData_List_Tgt[tgt_data]

                    if src_recursive_Data.IsSame(DataColl_This, DataColl_Tgt, dst_recursive_Data) == False:
                        continue

                    bFind = True
                if bFind == False:
                    return False
                else:
                    return True

            elif type(CheckData) != type(False):
                return False

            if CheckData == False:
                return True

        if not PyUtils.IsSameValue(src_data, tgt_data):
            return False

        return True

    def IsSame(self, DataColl_This, DataColl_Tgt, TgtDataD):
        if self.strID != TgtDataD.strID:
            return False

        #checker는 Collector 둘중 아무거나에서 가져와도 됨
        CheckerColl = DataColl_This
        if self.strID not in CheckerColl.CheckerList:
            return False

        list_cnt = len(self.list)
        if list_cnt != len(TgtDataD.list):
            return False
        
        CheckerD = CheckerColl.CheckerList[self.strID]

        if CheckerD.bTemp == False:
            if list_cnt != len(CheckerD.list):
                return False

        #with concurrent.futures.ThreadPoolExecutor(max_workers=8) as executor:
        #    futures = {executor.submit(self.IsSameElement, DataColl_This, DataColl_Tgt, CheckerD.list[i] if i < len(CheckerD.list) else True , self.list[i], TgtDataD.list[i], i < len(CheckerD.list)): i for i in range(list_cnt)}
        #    for future in concurrent.futures.as_completed(futures):
        #        if not future.result():
        #            return False

        for i in range(list_cnt):
            if not self.IsSameElement(DataColl_This, DataColl_Tgt, CheckerD.list[i] if i < len(CheckerD.list) else True , self.list[i], TgtDataD.list[i], i < len(CheckerD.list)):
                return False
                
        return True

class CheckerCollection:
    def __init__(self):
        self.FileFrom = ''
        self.CheckerList = {}
        self.DataKeyList = []
        self.DataList = {}
        self.__PrevID = ''

    def CopyFrom(self, CheckColl):
        self.CheckerList = copy.deepcopy(CheckColl.CheckerList)
        self.DataList = copy.deepcopy(CheckColl.DataList)

    def MakeChekcerListFromExcel(self, CheckerExcelPath):
        wb = openpyxl.load_workbook(CheckerExcelPath)
        sheet = wb.active

        nRow = 2
        while True:
            CheckerD = Checker()
            nCol = 2

            # 설마 1만개 이상 나오겠음..? 무한루프 방지
            strID = sheet.cell(row=nRow, column=1).value
            if strID == 'END' or nRow >= 10000:
                break

            if strID == None or strID == '':
                strID = self.__PrevID
                CheckerD = self.CheckerList[strID]
                CheckerD.list.append(False)
            else:
                self.__PrevID = strID

            CheckerD.strID = strID

            while True:
                value = sheet.cell(row=nRow, column=nCol).value

                # 설마 1천개 이상 있겠음..? 무한루프 방지
                if value == 'END' or nCol >= 1000:
                    break

                if value == 'TEMPEND':
                    CheckerD.bTemp = True
                    break

                if value == 'KEY':
                    CheckerD.KeyIndex = nCol-2
                    value = False

                if type(value) == type(''):
                    if value.find(',') != -1:
                        value = value.split(',')

                CheckerD.list.append(value)
                nCol += 1

            self.CheckerList[strID] = CheckerD
            nRow += 1

    def ExistChecker(self, strID):
        return strID in self.CheckerList

    def GetCheckDataCount(self, strID):
        return len(self.CheckerList[strID].list)

    def MakeToChecker(self, data_list) -> Checker():
        strID = data_list[0]

        if type(strID) != type(''):
            return False

        CheckerD = Checker()
        CheckerD.strID = strID

        for strData in data_list[1:]:
            if type(strData) == type(''):
                if PyUtils.IsSimmilarByPattern(strData, '[+-]{0,1}\\d+\\.\\d*') == True:
                    CheckerD.list.append(float(strData))
                elif PyUtils.IsSimmilarByPattern(strData, '\\d+') == True:
                    CheckerD.list.append(int(strData))
                else:
                    CheckerD.list.append(strData)
            else:
                CheckerD.list.append(strData)

        return CheckerD

    def AddData(self, data_list, nLine):
        DataD = self.MakeToChecker(data_list)
        DataD.nLine = nLine+1

        strID = DataD.strID

        KeyIndex = self.CheckerList[strID].KeyIndex
        
        if strID not in self.DataKeyList:
            self.DataKeyList.append(strID)
            
        if KeyIndex == -1:
            if strID not in self.DataList:
                self.DataList[strID] = []
            self.DataList[strID].append(DataD)
        else:
            if strID not in self.DataList:
                self.DataList[strID] = {}
            self.DataList[strID][DataD.list[KeyIndex]] = DataD

        return True

    def AddTempChecker(self, strID):
        # 일단 문자+숫자로 이루어진 문자열만 strID로 쓰도록 함. 설마 이렇게 생긴거 말고 있겠어..?
        if PyUtils.IsSimmilarByPattern(strID, '[\\w\\d]+') == False:
            return False

        if strID in self.CheckerList:
            return False

        data_list = []
        data_list.append(strID)
        data_list.append(False) #Key
        CheckerD = self.MakeToChecker(data_list)
        CheckerD.KeyIndex = 0
        CheckerD.bTemp = True

        self.CheckerList[strID] = CheckerD
        return True

    def IsValidData(self, parsed_list) -> bool:
        if not self.ExistChecker(parsed_list[0]):
            return False

        CheckerD = self.CheckerList[parsed_list[0]]
        if CheckerD.bTemp == False:
            return len(parsed_list[1:]) == len(CheckerD.list)
        else:
            return True

    def AddDataListFromMecFile(self, mec_path):
        self.FileFrom = mec_path
        mecfile = open(mec_path, "r")
        mecfile_lines = mecfile.readlines()

        pbar = tqdm(total=len(mecfile_lines), desc='Parsing...')

        i = 0
        while i < len(mecfile_lines):
            prev_line = i
            (parsed_list, i) = PyUtils.GetParsedBlock(mecfile_lines, i)

            pbar.update(i-prev_line)

            if len(parsed_list) < 2:
                continue
            
            if self.IsValidData(parsed_list):
                self.AddData(parsed_list, prev_line)
            else:
                if self.AddTempChecker(parsed_list[0]):
                    self.AddData(parsed_list, prev_line)

        pbar.close()